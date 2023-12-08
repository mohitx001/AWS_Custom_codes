import boto3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import argparse

# Initialize the argument parser
parser = argparse.ArgumentParser(description='Extract instance details and save to Excel')
parser.add_argument('-r', '--region', help='AWS Region', required=True)
args = parser.parse_args()

# Initialize the Boto3 client with the specified region
ec2 = boto3.client('ec2', region_name=args.region)
autoscaling = boto3.client('autoscaling', region_name=args.region)  # New line to initialize Auto Scaling client

# Get all instances
instances = ec2.describe_instances()

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Instance Details'

# Write account ID
account_id = boto3.client('sts').get_caller_identity().get('Account')
account_cell = sheet['A1']
account_cell.value = f'Account ID: {account_id}'
account_cell.font = Font(bold=True)
account_cell.alignment = Alignment(horizontal='center')
account_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
sheet.merge_cells('A1:L1')

# Write headers
headers = ['Service', 'Instance_Name', 'Instance_ID', 'Instance Class', 'State', 'Platform', 'Region', 'Monitoring', 'SSM_Agent', 'CW_Agent', 'AutoScalingGroup', 'Remarks']  # Updated headers
for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write instance details to Excel
row = 3
for reservation in instances['Reservations']:
    for instance in reservation['Instances']:
        instance_id = instance['InstanceId']
        instance_name = ''
        for tag in instance['Tags']:
            if tag['Key'] == 'Name':
                instance_name = tag['Value']
                break
        instance_class = instance['InstanceType']
        state = instance['State']['Name']
        platform = instance.get('Platform', 'Linux/Unix')  # Default to Linux/Unix if platform is not specified
        region = args.region
        monitoring_state = instance['Monitoring']['State'] if 'Monitoring' in instance and 'State' in instance['Monitoring'] else 'Disabled'

        # Check if the instance is part of an Auto Scaling group
        response = autoscaling.describe_auto_scaling_instances(InstanceIds=[instance_id])
        auto_scaling_group = response['AutoScalingInstances'][0]['AutoScalingGroupName'] if response['AutoScalingInstances'] else 'Not in Auto Scaling Group'

        data = ['EC2', instance_name, instance_id, instance_class, state, platform, region, monitoring_state, '', '', auto_scaling_group, '']  # Updated data
        for col, value in enumerate(data, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            if row > 2 or col > 1:  # Align everything to the left except for row 1, row 2, and column 1
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # Keep row 1, row 2, and column 1 centered
                cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

# Merge the cells in the first column for the "EC2" service
service_cell = sheet.cell(row=3, column=1)
last_row = row - 1
sheet.merge_cells(start_row=3, start_column=1, end_row=last_row, end_column=1)
service_cell.value = 'EC2'
service_cell.font = Font(bold=True)
service_cell.alignment = Alignment(horizontal='center', vertical='center')
service_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

# Save the Excel workbook
workbook.save(f'all_instance_details_{account_id}_{args.region}.xlsx')
