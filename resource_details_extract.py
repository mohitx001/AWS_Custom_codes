#testing
import boto3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import argparse

# Initialize the argument parser
parser = argparse.ArgumentParser(description='Extract instance details and save to Excel')
parser.add_argument('-r', '--region', help='AWS Region', required=True)
args = parser.parse_args()

# Initialize the Boto3 client with the specified region
ec2 = boto3.client('ec2', region_name=args.region)
autoscaling = boto3.client('autoscaling', region_name=args.region)  # New line to initialize Auto Scaling client
# Get all instances
instances = ec2.describe_instances()

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Instance Details'

# Write account ID
account_id = boto3.client('sts').get_caller_identity().get('Account')
account_cell = sheet['A1']
account_cell.value = f'Account ID: {account_id} ({args.region})'
account_cell.font = Font(bold=True)
account_cell.alignment = Alignment(horizontal='center')
account_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
sheet.merge_cells('A1:K1')

# Write headers
headers = ['Service', 'Instance_Name', 'Instance_ID', 'Instance Class', 'State', 'Platform', 'Monitoring', 'SSM_Agent', 'CW_Agent', 'AutoScalingGroup', 'Remarks']  # Updated headers
for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write EC2 details to Excel
row = 3
for reservation in instances['Reservations']:
    for instance in reservation['Instances']:
        instance_id = instance['InstanceId']
        instance_name = ''
        for tag in instance['Tags']:
            if tag['Key'] == 'Name':
                instance_name = tag['Value']
                break
        instance_class = instance['InstanceType']
        state = instance['State']['Name']
        platform = instance.get('Platform', 'Linux/Unix')  # Default to Linux/Unix if platform is not specified
        region = args.region
        monitoring_state = instance['Monitoring']['State'] if 'Monitoring' in instance and 'State' in instance['Monitoring'] else 'Disabled'

        # Check if the instance is part of an Auto Scaling group
        response = autoscaling.describe_auto_scaling_instances(InstanceIds=[instance_id])
        auto_scaling_group = response['AutoScalingInstances'][0]['AutoScalingGroupName'] if response['AutoScalingInstances'] else 'Not in Auto Scaling Group'

        data = ['EC2', instance_name, instance_id, instance_class, state, platform, monitoring_state, '', '', auto_scaling_group, '']  # Updated data
        for col, value in enumerate(data, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            if row > 2 or col > 1:  # Align everything to the left except for row 1, row 2, and column 1
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # Keep row 1, row 2, and column 1 centered
                cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

# Merge the cells in the first column for the "EC2" service
service_cell = sheet.cell(row=3, column=1)
last_row = row - 1
sheet.merge_cells(start_row=3, start_column=1, end_row=last_row, end_column=1)
service_cell.value = 'EC2'
service_cell.font = Font(bold=True)
service_cell.alignment = Alignment(horizontal='center', vertical='center')
service_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

# Get all RDS instances
rds = boto3.client('rds', region_name=args.region)
rds_instances = rds.describe_db_instances()
# Write RDS headers
rds_headers = ['RDSname', 'RDS Endpoint', 'Class', 'State', 'Engine', 'Version', 'Remarks']
for col, header in enumerate(rds_headers, start=2):
    cell = sheet.cell(row=row, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write RDS details to Excel
row += 1
#row is 9 here
for rds_instance in rds_instances['DBInstances']:
    rds_name = rds_instance['DBInstanceIdentifier']
    rds_endpoint = rds_instance['Endpoint']['Address']
    rds_class = rds_instance['DBInstanceClass']
    rds_state = rds_instance['DBInstanceStatus']
    rds_engine = rds_instance['Engine']
    rds_version = rds_instance['EngineVersion']
    rds_remarks = ''  # You can add remarks here if needed

    rds_data = [rds_name, rds_endpoint, rds_class, rds_state, rds_engine, rds_version, rds_remarks]
    for col, value in enumerate(rds_data, start=2):
        cell = sheet.cell(row=row, column=col, value=value)
        if row > 2 or col > 1:  # Align everything to the left except for row 1, row 2, and column 1
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:  # Keep row 1, row 2, and column 1 centered
            cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
# Merge the cells in the first column for the "RDS" service
# row is 11 here
rds_service_cell = sheet.cell(row=last_row+1, column=1)
rds_last_row = row - 1 # 10
sheet.merge_cells(start_row=last_row+1, start_column=1, end_row=rds_last_row, end_column=1)
sheet.merge_cells(start_row=last_row+1, start_column=8, end_row=last_row+1, end_column=11)
rds_service_cell.value = 'RDS'
rds_service_cell.font = Font(bold=True)
rds_service_cell.alignment = Alignment(horizontal='center', vertical='center')
rds_service_cell.fill = PatternFill(start_color="FFFAF0", end_color="FFFAF0", fill_type="solid")

# Get all Auto Scaling groups
autoscaling_groups = autoscaling.describe_auto_scaling_groups()

# Write Auto Scaling Group headers
asg_headers = ['ASG Name', 'Launch Configuration', 'Launch Template' , 'Min Size', 'Max Size', 'Desired Capacity' ,'Remarks' ]
for col, header in enumerate(asg_headers, start=2):
    cell = sheet.cell(row=row, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write Auto Scaling Group details to Excel
row += 1
for group in autoscaling_groups['AutoScalingGroups']:
    asg_name = group['AutoScalingGroupName']
    launch_config = group.get('LaunchConfigurationName', 'N/A')  # Check if key is present
    #Launch_template =  group['LaunchTemplate'].get('LaunchTemplateName', 'N/A')
    Launch_template = group.get('LaunchTemplate', {}).get('LaunchTemplateName', 'N/A')
    min_size = group['MinSize']
    max_size = group['MaxSize']
    desired_capacity = group['DesiredCapacity']

    asg_data = [asg_name, launch_config,Launch_template, min_size, max_size, desired_capacity,'']
    for col, value in enumerate(asg_data, start=2):
        cell = sheet.cell(row=row, column=col, value=value)
        if row > 2 or col > 1:  # Align everything to the left except for row 1, row 2, and column 1
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:  # Keep row 1, row 2, and column 1 centered
            cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

# Merge the cells in the first column for the "ASG" service
asg_service_cell = sheet.cell(row=rds_last_row+1, column=1)
asg_last_row = row - 1
sheet.merge_cells(start_row=rds_last_row+1, start_column=1, end_row=asg_last_row, end_column=1)
sheet.merge_cells(start_row=rds_last_row+1, start_column=8, end_row=rds_last_row+1, end_column=11)
asg_service_cell.value = 'ASG'
asg_service_cell.font = Font(bold=True)
asg_service_cell.alignment = Alignment(horizontal='center', vertical='center')
asg_service_cell.fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")

# Save the Excel workbook

workbook.save(f'all_instance_details_{account_id}_{args.region}.xlsx')
