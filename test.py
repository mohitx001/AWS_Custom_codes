import boto3
import openpyxl
import argparse

# Initialize the argument parser
parser = argparse.ArgumentParser(description='Extract instance details and save to Excel')
parser.add_argument('-r', '--region', help='AWS Region', required=True)
args = parser.parse_args()

# Create an EC2 client
ec2_client = boto3.client('ec2', region_name=args.region)

# Describe EC2 instances
instances = ec2_client.describe_instances()

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Agent Status'

# Add headers to the Excel sheet
sheet['A1'] = 'Instance ID'
sheet['B1'] = 'SSM Agent Status'
sheet['C1'] = 'CloudWatch Agent Status'

row = 2  # Start from the second row for data

# Iterate through each instance
for reservation in instances['Reservations']:
    for instance in reservation['Instances']:
        instance_id = instance['InstanceId']
        
        # Check if SSM agent is installed
        ssm_client = boto3.client('ssm', region_name=args.region)
        ssm_status = ''
        try:
            ssm_response = ssm_client.describe_instance_information(InstanceInformationFilterList=[{'key': 'InstanceIds', 'valueSet': [instance_id]}])
            if ssm_response['InstanceInformationList']:
                ssm_status = 'Installed'
            else:
                ssm_status = 'Not Installed'
        except ssm_client.exceptions.InvalidInstanceId:
            ssm_status = 'Invalid Instance ID or Not Installed'
        
        # Check if CloudWatch agent is installed
        cw_client = boto3.client('cloudwatch', region_name=args.region)
        cw_status = 'Not Installed'
        # List of metrics to check for CloudWatch agent
        metrics_to_check = ['MemoryUtilization', 'DiskUtilization']
        try:
            cw_response = cw_client.describe_alarms_for_metric(
                MetricName='mem_used_percent',
                Namespace='System/Linux',
                Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}]
            )
            if cw_response['MetricAlarms']:
                cw_status = 'Installed'
            else:
                cw_status = 'Not Installed'
        except cw_client.exceptions.ResourceNotFoundException:
            cw_status = 'Not Installed or No Alarms Configured'
        
        # Write the results to the Excel sheet
        sheet.cell(row=row, column=1, value=instance_id)
        sheet.cell(row=row, column=2, value=ssm_status)
        sheet.cell(row=row, column=3, value=cw_status)
        row += 1

# Save the workbook to a file
workbook.save('agent_status.xlsx')