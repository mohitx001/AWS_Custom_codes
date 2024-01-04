import boto3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import argparse

# Initialize the argument parser
parser = argparse.ArgumentParser(description='Extract Auto Scaling Group details and save to Excel')
parser.add_argument('-r', '--region', help='AWS Region', required=True)
args = parser.parse_args()

# Initialize the Boto3 client for Auto Scaling with the specified region
autoscaling = boto3.client('autoscaling', region_name=args.region)

# Get all Auto Scaling groups
autoscaling_groups = autoscaling.describe_auto_scaling_groups()

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'AutoScaling Details'

# Write headers
headers = ['AutoScaling Group Name', 'Launch Configuration', 'Launch Template', 'Min Size', 'Max Size', 'Desired Capacity']
for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write Auto Scaling group details to Excel
row = 2
for group in autoscaling_groups['AutoScalingGroups']:
    group_name = group['AutoScalingGroupName']
    launch_config = group.get('LaunchConfigurationName', 'N/A')  # Check if key is present
    #Launch_template =  group['LaunchTemplate'].get('LaunchTemplateName', 'N/A')
    Launch_template = group.get('LaunchTemplate', {}).get('LaunchTemplateName', 'N/A')  # Check if key is present
    min_size = group['MinSize']
    max_size = group['MaxSize']
    desired_capacity = group['DesiredCapacity']

    data = [group_name, launch_config, Launch_template, min_size, max_size, desired_capacity]
    for col, value in enumerate(data, start=1):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

# Save the Excel workbook
workbook.save(f'all_autoscaling_details_{args.region}.xlsx')
