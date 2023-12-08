import boto3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import argparse

# Initialize the argument parser
parser = argparse.ArgumentParser(description='Extract RDS details and save to Excel')
parser.add_argument('-r', '--region', help='AWS Region', required=True)
args = parser.parse_args()

# Initialize the Boto3 client with the specified region
rds = boto3.client('rds', region_name=args.region)

# Get all RDS instances
rds_instances = rds.describe_db_instances()

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'RDS Details'

# Write account ID
account_id = boto3.client('sts').get_caller_identity().get('Account')
account_cell = sheet['A1']
account_cell.value = f'Account ID: {account_id}'
account_cell.font = Font(bold=True)
account_cell.alignment = Alignment(horizontal='center')
account_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
sheet.merge_cells('A1:H1')

# Write headers
headers = ['Service', 'RDS Name', 'RDS Endpoint', 'Class', 'State', 'Engine', 'Version', 'Remark']
for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write RDS details to Excel
row = 3
for rds_instance in rds_instances['DBInstances']:
    rds_name = rds_instance['DBInstanceIdentifier']
    rds_endpoint = rds_instance['Endpoint']['Address']
    rds_class = rds_instance['DBInstanceClass']
    rds_state = rds_instance['DBInstanceStatus']
    rds_engine = rds_instance['Engine']
    rds_version = rds_instance['EngineVersion']

    data = ['RDS', rds_name, rds_endpoint, rds_class, rds_state, rds_engine, rds_version, '']
    for col, value in enumerate(data, start=1):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

# Merge the cells in the first column for the "RDS" service
for i in range(3, row):
    sheet.cell(row=i, column=1).value = 'RDS'
    sheet.cell(row=i, column=1).font = Font(bold=True)
    sheet.cell(row=i, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i, column=1).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

# Save the Excel workbook
workbook.save(f'all_rds_details_{args.region}.xlsx')
