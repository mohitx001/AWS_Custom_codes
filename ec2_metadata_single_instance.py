import boto3
import openpyxl
import argparse

# Initialize the argument parser
parser = argparse.ArgumentParser(description='Extract instance details and save to Excel')
parser.add_argument('-i', '--instance-id', help='Instance ID', required=True)
parser.add_argument('-r', '--region', help='AWS Region', required=True)
args = parser.parse_args()

# Initialize the Boto3 client with the specified region
ec2 = boto3.client('ec2', region_name=args.region)

# Get instance details
instance = ec2.describe_instances(InstanceIds=[args.instance_id])

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Instance Details'

# Write headers
headers = ['Instance Name', 'Instance Class', 'Key Name', 'Security Group Name', 'Subnet ID', 'Volume', 'Platform']
for col, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col, value=header)

# Write instance details to Excel
instance = instance['Reservations'][0]['Instances'][0]
instance_name = ''
for tag in instance['Tags']:
    if tag['Key'] == 'Name':
        instance_name = tag['Value']
        break
instance_class = instance['InstanceType']
key_name = instance.get('KeyName', 'N/A')
security_group_name = ', '.join([group['GroupName'] for group in instance['SecurityGroups']])
subnet_id = instance['SubnetId']
volume = ', '.join([vol['Ebs']['VolumeId'] for vol in instance['BlockDeviceMappings']])
platform = instance.get('Platform', 'Linux/Unix')  # Default to Linux/Unix if platform is not specified

data = [instance_name, instance_class, key_name, security_group_name, subnet_id, volume, platform]
for col, value in enumerate(data, start=1):
    sheet.cell(row=2, column=col, value=value)

# Save the workbook
workbook.save(f'instance_{args.instance_id}_details.xlsx')
